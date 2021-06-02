import re
from typing import Optional

from openpyxl.worksheet.worksheet import Worksheet
from spellchecker import SpellChecker

from Utils import Utils


# noinspection PyMethodMayBeStatic
class AbbreviationFixer:

    def __init__(self, spell_checker: SpellChecker):
        self.spell_checker = spell_checker
        self.spell_checker.word_frequency.load_dictionary('Files/dict.json', encoding='utf-8')
        self.abbr_checker = SpellChecker(language='ru', local_dictionary='Files/dict.json')
        """
        key - номер столбца в таблице
        value - массив из трех значений. [кол-во сокращений, кол-во нуллов, кол-во строк]
        int: list[int]
        """
        self.abbr_counter = {}
        self.abbr_storage = {}

    def correct_abbreviation(self, string: str) -> str:
        """
        Исправляет аббревиатуры
        :param string: значение ячейки
        :return: новое значение ячейки
        """
        new_value = ''
        values = re.split('(\d+)', string)
        for value in values:
            if value.isnumeric():
                new_value += value
                continue

            if value == '':
                continue

            new_value += self.spell_checker.correction(value)
        return new_value

    def update_abbr_storage(self, value: Optional[str], index: int) -> None:
        if index not in self.abbr_storage:
            self.abbr_storage[index] = []

        self.abbr_storage[index].append(value)

    def _is_abbr_last_in_cell(self, cell_value: str, column_index: int) -> bool:
        """
        проверяет, стоит ли сокращение последним в ячейке
        :param cell_value: значение ячейки
        :return: True, если сокращение - последнее. Иначе False
        """
        values = re.split('(\d+)', cell_value)
        pseudo_last_value = values[-1].strip()
        if column_index not in self.abbr_storage:
            self.abbr_storage[column_index] = []
        if pseudo_last_value == '':
            self.update_abbr_storage(None, column_index)
            return False

        last_value = pseudo_last_value.split(' ')[-1]

        if len(self.abbr_checker.known([last_value])) > 0:
            self.update_abbr_storage(last_value, column_index)
            return True

        self.update_abbr_storage(None, column_index)
        return False

    def update_abbr_counter(self, cell_value: str, column_index: int, row_counter: int) -> None:
        """
        обновляет данные по столбцу
        :param row_counter: кол-во строк
        :param cell_value: значение ячейки
        :param column_index: номер столбца
        :return: None
        """
        if column_index not in self.abbr_counter:
            self.abbr_counter[column_index] = [0, 0, row_counter]

        if cell_value is None:
            self.abbr_counter[column_index][1] += 1
            self.update_abbr_storage(None, column_index)
            return

        if self._is_abbr_last_in_cell(cell_value, column_index):
            self.abbr_counter[column_index][0] += 1

    def create_new_value_for_old_cell(self, old_value: str, new_value: str) -> str:
        if new_value is None:
            return old_value
        return_value = old_value
        try:
            return_value = ''.join(old_value.rsplit(new_value, 1))

        except AttributeError:
            #?????????????????
            print(new_value)
        finally:
            return return_value

    def create_cols_for_abbrs(self, sheet: Worksheet) -> None:
        for key, value in self.abbr_counter.items():
            clean_rows = value[2] - value[1]
            if value[0] * 2 < clean_rows:
                continue

            sheet.insert_cols(key + 1)
            abbrs = self.abbr_storage[key]
            for i in range(2, sheet.max_row + 1):
                cell = sheet.cell(i, key + 1)
                cell.border = Utils.get_thin_border()
                new_value = 'NA' if abbrs[i - 2] is None else abbrs[i - 2]
                cell.value = new_value
                old_cell = sheet.cell(i, key)
                old_cell.value = self.create_new_value_for_old_cell(old_cell.value, abbrs[i - 2])
