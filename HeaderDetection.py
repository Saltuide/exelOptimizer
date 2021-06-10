from re import sub

from openpyxl.worksheet.worksheet import Worksheet
from wiki_ru_wordnet import WikiWordnet


class HeaderDetection:

    def __init__(self, sheet: Worksheet):
        self.sheet = sheet
        self.header_end = 0
        self.hypernyms = {}
        self.header = ''

    def create_header(self):
        wiki_wordnet = WikiWordnet()

        for i in range(1, self.sheet.max_row):
            border_counter = 0
            for j in range(1, self.sheet.max_column):
                _cell_value = self.sheet.cell(i, j).value

                if not _cell_value:
                    continue

                _cell_value = sub('\W', '', _cell_value)
                if len(_cell_value) == 0:
                    continue

                values = sub('\W', ' ', self.sheet.cell(i, j).value).split()
                for value in values:
                    synsets = wiki_wordnet.get_synsets(value)
                    if not synsets:
                        continue
                    for synset in synsets:
                        for hypernym in wiki_wordnet.get_hypernyms(synset):
                            for w in hypernym.get_words():
                                if w.lemma() not in self.hypernyms:
                                    self.hypernyms[w.lemma()] = 1
                                else:
                                    self.hypernyms[w.lemma()] += 1

        self.header = max(self.hypernyms, key=self.hypernyms.get)

    def find_header_2(self):
        for i in range(1, self.sheet.max_row):
            border_counter = 0
            for j in range(1, self.sheet.max_column):
                _cell_value = self.sheet.cell(i, j).value
                if not _cell_value:
                    continue

                _cell_value = sub('\W', '', _cell_value)
                if len(_cell_value) > 0:
                    border_counter += 1

            if border_counter >= round(self.sheet.max_column * 0.8):
                self.header_end = i - 2
                break
            border_counter = 0

        if self.header_end == 0:
            self.create_header()

    def find_header(self):
        for i in range(1, self.sheet.max_row):
            border_counter = 0
            for j in range(1, self.sheet.max_column):
                _cell = self.sheet.cell(i, j)
                if _cell.border.bottom.style is not None:
                    border_counter += 1

            if border_counter > self.sheet.max_column // 2:
                self.header_end = i - 1
                break
            border_counter = 0

        if self.header_end == 0:
            self.find_header_2()
