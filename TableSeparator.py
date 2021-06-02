import pandas as pd
import numpy as np
from typing import NoReturn
from copy import copy

import openpyxl
from skimage.measure import label, regionprops
from openpyxl.worksheet.worksheet import Worksheet, Cell
from openpyxl.utils import get_column_letter


class TableSeparator:

    def __init__(self, file: str):
        self.file = file
        """
        {
            table_number: {
                rows:{ 
                    start: index
                    end: index
                    },
                cols:{ 
                    start: index
                    end: index
                    }
                }
        }
        """
        self.table_borders = {}

    def open_file(self):
        try:
            self.df = pd.read_excel('Files/main.xlsx', index_col=None, header=None)
        except FileNotFoundError:
            print("Не удалось открыть файл")

    def _create_element_for_table_borders(self, index: int, col_indexes, row_indexes):
        row_borders = {'start': row_indexes[0] + 1, 'end': row_indexes[-1] + 2}
        col_borders = {'start': col_indexes[0] + 1, 'end': col_indexes[-1] + 2}
        self.table_borders[index] = {'rows': row_borders, 'cols': col_borders}

    def get_tables_borders(self):
        self.open_file()

        binary_representation = np.array(self.df.notnull().astype('int'))

        list_of_dataframes = []
        connected_df = label(binary_representation)
        for s in regionprops(connected_df):
            list_of_dataframes.append(self.df.iloc[s.bbox[0]:s.bbox[2], s.bbox[1]:s.bbox[3]])

        for i, df in enumerate(list_of_dataframes):
            cols_index = df.columns.to_numpy()
            rows_index = df.index.to_numpy()
            self._create_element_for_table_borders(i, cols_index, rows_index)

    def _copy_cell(self, old_cell: Cell, new_cell: Cell) -> NoReturn:
        new_cell.value = old_cell.value
        new_cell.style = old_cell.style
        new_cell.border = copy(old_cell.border)

    def _set_proper_cols_width(self, sheet: Worksheet) -> NoReturn:
        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value + 5

    def _transfer_data_to_new_sheet(self, old_sheet: Worksheet, new_sheet: Worksheet, borders: dict):
        rows = borders['rows']
        cols = borders['cols']

        row_counter = 1
        for row in range(rows['start'], rows['end']):
            col_counter = 1
            for col in range(cols['start'], cols['end']):
                self._copy_cell(old_sheet.cell(row, col), new_sheet.cell(row_counter, col_counter))
                col_counter += 1

            row_counter += 1

    def separate_tables_to_different_worksheets(self):
        if len(self.table_borders) < 2:
            return
        wb = openpyxl.load_workbook(self.file, data_only=True)
        active_sheet = wb.active
        for table_index, borders in self.table_borders.items():
            title = f'New Worksheet number {table_index + 1}'
            wb.create_sheet(title)
            new_sheet = wb[title]
            self._transfer_data_to_new_sheet(active_sheet, new_sheet, borders)
            self._set_proper_cols_width(new_sheet)

        wb.save('Files/test.xlsx')
