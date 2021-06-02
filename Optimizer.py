import re

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet, Cell
from spellchecker import SpellChecker

from AbbreviationFixer import AbbreviationFixer

UNDEFINED_CELL = 'NA'


# noinspection PyMethodMayBeStatic
class Optimizer:

    def __init__(self, file: Workbook):
        self.file = file
        self.spell_checker = SpellChecker(language='ru')
        self.abbreviation_fixer = AbbreviationFixer(self.spell_checker)

    def strip(self, cell: Cell) -> None:
        """
        Убирает пробелы по краям

        :param cell: Ячейка для обработки
        :return: None
        """
        if isinstance(cell.value, str):
            value = cell.value
            cell.value = value.strip()

    def set_cell_not_null(self, cell: Cell) -> None:
        """
        Присваивает обозначение для пустой ячейки

        :param cell: Пустая ячейка
        :return: None
        """
        cell.value = UNDEFINED_CELL

    def is_string(self, value: str) -> bool:
        """
        Проверяет, является ли значение ячейки строкой

        :param value: Строка для проверки
        :return: True, если это строка
        """
        if value.isalpha():
            return True
        return False

    def check_spelling(self, value: str) -> str:
        """
        Исправляет синтаксические ошибки в строке

        :param value: Строка для исправления ошибок
        :return: Исправленная строка
        """
        if len(value) < 3:
            return value

        return self.spell_checker.correction(value)

    def create_new_cell_value(self, strings: list, delimiters: str) -> str:
        """
        Собирает новое исправленное значение ячейки

        :param strings: Массив строковых значений
        :param delimiters: Строка с исходными разделителями
        :return: Новое значение ячейки
        """
        new_value = ''
        for counter, string in enumerate(strings):
            if self.is_string(string):
                new_value += self.check_spelling(string)
                if counter < len(delimiters):
                    new_value += delimiters[counter]
            else:
                new_value += string
                if counter < len(delimiters):
                    new_value += delimiters[counter]

        return new_value

    def sheet_spelling_lopper(self, sheet: Worksheet) -> None:
        """
        Проходит по листу и исправляет базовые ошибки

        :param sheet: Текущий лист
        :return: None
        """
        for i in range(2, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=i, column=j)

                self.strip(cell)

                if cell.value is None or cell.value == '':
                    self.set_cell_not_null(cell)
                    self.abbreviation_fixer.update_abbr_counter(None, j, sheet.max_column)
                    continue

                try:
                    value = cell.value.replace('_', '$%%$')
                    strings = re.sub('\W', ' ', value).split()
                    delimiters = re.sub('\w', '', value).replace('$%%$', '_')
                except Exception:
                    self.abbreviation_fixer.update_abbr_storage(None, j)
                    continue

                new_cell_value = self.create_new_cell_value(strings, delimiters)
                cell.value = self.abbreviation_fixer.correct_abbreviation(new_cell_value)
                self.abbreviation_fixer.update_abbr_counter(cell.value, j, sheet.max_column)
        self.abbreviation_fixer.create_cols_for_abbrs(sheet)

    def a(self):
        for sheet in self.file.worksheets:
            self.sheet_spelling_lopper(sheet)
            self.file.save('Files/new-main.xlsx')
